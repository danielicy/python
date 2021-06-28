import openpyxl


def iterateAllCells(sheet):
    for row in range(1, sheet.max_row + 1):
        for column in range(1, sheet.max_column + 1):
            cell = sheet.cell(row, column)
            print(cell.value)


#wb = openpyxl.Workbook()
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)
sheet = wb["Sheet1"]
cell = sheet["a1"]
column = sheet["a"]
cells = sheet["a:c"]
sheet["a1:c3"]
sheet[1:3]
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)


cell2 = sheet.cell(row=1, column=1)
print(sheet.max_column)

print(sheet.max_row)

iterateAllCells(sheet)
